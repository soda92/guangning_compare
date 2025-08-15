# -*- coding: gbk -*-

import csv
import os
from openpyxl import load_workbook, Workbook


def excel_append2(file_path, column_headers, contents):
    # ����������ݳ���һ����
    if len(column_headers) != len(contents):
        raise ValueError("The number of column headers must match the number of contents.")

    file_extension = os.path.splitext(file_path)[1].lower()

    if file_extension == '.xlsx':
        try:
            # ���Լ������е�Excel�ļ�
            workbook = load_workbook(file_path)
            sheet = workbook.active
        except FileNotFoundError:
            # �ļ������ڣ��������ļ���д���ͷ
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(column_headers)  # д���ͷ

        # ��������д���µ�һ��
        sheet.append(contents)
        workbook.save(file_path)
        print(f"�ѽ�����׷���� {file_path}")

    elif file_extension == '.csv':
        # ���CSV�ļ��Ƿ��Ѵ���
        file_exists = os.path.exists(file_path)

        with open(file_path, mode='a', encoding='utf-8', newline='') as file:
            writer = csv.writer(file)

            if not file_exists:
                # ����ļ������ڣ�д���ͷ
                writer.writerow(column_headers)

            # д������
            writer.writerow(contents)
        print(f"�ѽ�����׷���� {file_path}")

    else:
        raise ValueError("Unsupported file format. Only .xlsx and .csv are supported.")
